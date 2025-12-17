from __future__ import annotations

import json
import logging
from typing import List, Optional
from pathlib import Path
import uuid

import re

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

from ...domain.invoice import Invoice
from ...services.ocr_service import OcrService
from ...services.excel_service import ExcelService
from ...services.pdf_text_service import extract_text_from_pdf_bytes
from ...config import load_app_config

# ロガー設定
logger = logging.getLogger(__name__)

router = APIRouter()

# job_id -> Excel file path
_excel_jobs: dict[str, str] = {}


def _japanese_ratio(text: str) -> float:
    if not text:
        return 0.0
    japanese_chars = re.findall(r"[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]", text)
    total_chars = len(re.sub(r"\s+", "", text))
    return (len(japanese_chars) / total_chars) if total_chars else 0.0

# リクエストボディのモデル定義
class ExcelGenerationRequest(BaseModel):
    corp_name: str
    invoices_data: List[dict]  # OCR結果のリスト

@router.post("/process")
async def process_pdfs(
    corp_name: str = Form(""),
    address: str = Form(""),
    corp_number: str = Form(""),
    mode: str = Form(...),
    start_month: Optional[int] = Form(None),
    month_order: Optional[str] = Form("ascending"),
    month_mappings: str = Form(...),
    files: List[UploadFile] = File(...)
):
    """
    PDFファイルをOCR処理してExcelに反映する
    
    Args:
        corp_name: 法人名
        mode: "single" or "multi"
        start_month: 複数月モードの開始月
        month_order: 月の並び順 "ascending" (昇順) or "descending" (降順)
        month_mappings: JSON文字列 [{"filename": "xxx.pdf", "selectedMonth": 1}, ...]
        files: PDFファイルリスト
    
    Returns:
        処理結果のJSON
    """
    try:
        # 設定読み込み
        cfg = load_app_config()
        ocr_service = OcrService(cfg)
        excel_service = ExcelService(cfg)
        
        # 月マッピング情報をパース
        try:
            mappings = json.loads(month_mappings)
            month_map = {m['filename']: m['selectedMonth'] for m in mappings}
        except Exception as e:
            logger.error(f"月マッピングのパースに失敗: {e}")
            month_map = {}
        
        invoices: List[Invoice] = []
        results = []
        
        for file in files:
            try:
                # ファイル内容を読み込み
                content = await file.read()
                
                logger.info(f"処理開始: {file.filename}, mode={mode}")
                
                if mode == "single":
                    # 単月モード
                    selected_month = month_map.get(file.filename, 1)

                    # 1) まずOCR
                    invoice = ocr_service.analyze_invoice(
                        content,
                        mode="single",
                        start_month=None
                    )

                    ocr_confidence = invoice.fields.get("ocr_confidence", 0) if invoice.fields else 0
                    text_source = "ocr"
                    ocr_text = invoice.raw_text or ""
                    best_text = ocr_text

                    # OCRで既にkWhが取れているなら保持する（本文だけ差し替えるケースのため）
                    kwh_from_ocr = ""
                    if invoice.fields and selected_month:
                        kwh_from_ocr = invoice.fields.get(f"{selected_month}月値", "") or ""
                    if not kwh_from_ocr and selected_month and ocr_text:
                        kwh_from_ocr = OcrService._extract_kwh_from_text(ocr_text)

                    # 2) OCR品質が低い場合のみテキスト層へフォールバック
                    if best_text:
                        jr = _japanese_ratio(best_text)
                        should_fallback = (ocr_confidence < 0.8) or (jr < 0.2)
                    else:
                        should_fallback = True

                    if should_fallback:
                        extracted = extract_text_from_pdf_bytes(content)
                        if extracted:
                            best_text = extracted
                            text_source = "pdf_text"

                    # kWhは「OCRで取れていればそれを優先」。無い場合のみ、表示テキスト(best_text)から再抽出。
                    kwh_value = kwh_from_ocr
                    if not kwh_value and selected_month and best_text:
                        kwh_value = OcrService._extract_kwh_from_text(best_text)

                    if selected_month and kwh_value:
                        invoice.fields = {f"{selected_month}月値": kwh_value, "ocr_confidence": ocr_confidence}
                        logger.info(f"{file.filename}: {selected_month}月値={kwh_value}, 信頼度={ocr_confidence:.2f}")
                    else:
                        invoice.fields = {"ocr_confidence": ocr_confidence}
                        logger.warning(f"{file.filename}: kWh値を抽出できませんでした")

                    # 返却用テキストに合わせて raw_text も更新しておく（表示用途）
                    invoice.raw_text = best_text
                    
                    invoices.append(invoice)
                    results.append({
                        "filename": file.filename,
                        "status": "完了" if kwh_value else "kWh未検出",
                        "fields": invoice.fields,
                        "kwh": kwh_value,
                        "ocr_text": best_text,
                        "ocr_confidence": ocr_confidence
                        ,"text_source": text_source
                    })
                    
                else:
                    # 複数月モード

                    # 1) まずOCR
                    invoice = ocr_service.analyze_invoice(
                        content,
                        mode="multi",
                        start_month=start_month,
                        month_order=month_order
                    )

                    ocr_confidence = invoice.fields.get("ocr_confidence", 0) if invoice.fields else 0
                    text_source = "ocr"
                    best_text = invoice.raw_text or ""

                    # 2) OCR品質が低い場合のみテキスト層へフォールバック（表示用テキストの置き換え）
                    if best_text:
                        jr = _japanese_ratio(best_text)
                        should_fallback = (ocr_confidence < 0.8) or (jr < 0.2)
                    else:
                        should_fallback = True

                    if should_fallback:
                        extracted = extract_text_from_pdf_bytes(content)
                        if extracted:
                            best_text = extracted
                            text_source = "pdf_text"
                    
                    # kWh値が1つでも抽出できているか確認
                    kwh_extracted = any(key.endswith('月値') for key in invoice.fields.keys()) if invoice.fields else False
                    
                    invoices.append(invoice)
                    results.append({
                        "filename": file.filename,
                        "status": "完了" if kwh_extracted else "kWh未検出",
                        "fields": invoice.fields,
                        "ocr_text": best_text,
                        "ocr_confidence": ocr_confidence
                        ,"text_source": text_source
                    })
                
            except Exception as e:
                logger.error(f"{file.filename}の処理中にエラー: {str(e)}", exc_info=True)
                results.append({
                    "filename": file.filename,
                    "status": "エラー",
                    "error": str(e)
                })
        
        # Excelに書き込み
        try:
            excel_path = excel_service.write_invoices(
                invoices, 
                corp_name=corp_name,
                address=address,
                corp_number=corp_number
            )
            job_id = uuid.uuid4().hex
            _excel_jobs[job_id] = excel_path
            logger.info(f"Excel書き込み完了: {excel_path}")
        except Exception as e:
            logger.error(f"Excel書き込みエラー: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Excel書き込みエラー: {str(e)}")
        
        return JSONResponse({
            "success": True,
            "results": results,
            "excel_path": excel_path,
            "job_id": job_id
        })
        
    except Exception as e:
        logger.error(f"処理エラー: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download")
async def download_excel(
    job_id: str,
    corp_name: str = "",
    address: str = "",
    corp_number: str = "",
    kwh_overrides: str = ""
):
    """
    生成されたExcelファイルをダウンロード
    ダウンロード時に最新の住所・法人番号を更新
    """
    excel_path = _excel_jobs.get(job_id)
    if not excel_path or not Path(excel_path).exists():
        raise HTTPException(status_code=404, detail="Excelファイルが見つかりません（job_idが無効か期限切れです）")
    
    # 住所/法人番号/kWh上書きが指定されている場合、Excelファイルを更新
    if address or corp_number or kwh_overrides:
        try:
            from openpyxl import load_workbook
            cfg = load_app_config()
            wb = load_workbook(excel_path)
            sheet_name = cfg.get("excel_cell_map", {}).get("sheet", wb.sheetnames[0])
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 住所をB2に更新
            if address:
                ws['B2'] = address
            
            # 法人番号をB4に更新
            if corp_number:
                ws['B4'] = corp_number

            # kWhの上書き（{"1": "12345", "2": "23456" ...} のようなJSON文字列）
            if kwh_overrides:
                try:
                    overrides_raw = json.loads(kwh_overrides)
                except Exception as e:
                    logger.warning(f"kWh上書きのパースに失敗: {e}")
                    overrides_raw = {}

                month_cells = {
                    1: "B21",
                    2: "C21",
                    3: "D21",
                    4: "E21",
                    5: "F21",
                    6: "G21",
                    7: "H21",
                    8: "I21",
                    9: "J21",
                    10: "K21",
                    11: "L21",
                    12: "M21",
                }

                for k, v in (overrides_raw or {}).items():
                    try:
                        key_str = str(k).strip()
                        key_str = key_str.replace("月値", "").replace("月", "")
                        month = int(key_str)
                    except Exception:
                        continue
                    if month < 1 or month > 12:
                        continue

                    value_str = "" if v is None else str(v).strip()
                    if value_str == "":
                        continue

                    cell = month_cells.get(month)
                    if not cell:
                        continue

                    # 数値として書き込み（無理なら文字列）
                    try:
                        ws[cell] = int(value_str.replace(",", ""))
                    except Exception:
                        ws[cell] = value_str
            
            wb.save(excel_path)
            logger.info(f"Excelファイル更新: 住所={address}, 法人番号={corp_number}, kWh上書き={'あり' if kwh_overrides else 'なし'}")
        except Exception as e:
            logger.warning(f"Excelファイルの更新に失敗: {e}")
    
    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=Path(excel_path).name
    )


@router.post("/ocr_single")
async def process_single_pdf(
    mode: str = Form(...),
    selected_month: Optional[int] = Form(None),
    start_month: Optional[int] = Form(None),
    file: UploadFile = File(...)
):
    """
    単一のPDFファイルをOCR処理して結果を返す
    """
    try:
        # 設定読み込み
        cfg = load_app_config()
        ocr_service = OcrService(cfg)
        
        content = await file.read()
        logger.info(f"単一処理開始: {file.filename}, mode={mode}, month={selected_month}")
        
        invoice = None
        kwh_value = None
        
        if mode == "single":
            # 単月モード
            invoice = ocr_service.analyze_invoice(
                content,
                mode="single",
                start_month=None
            )
            
            # OCRテキストから直接kWh値を抽出
            if selected_month and invoice.raw_text:
                kwh_value = OcrService._extract_kwh_from_text(invoice.raw_text)
                if kwh_value:
                    invoice.fields = {f"{selected_month}月値": kwh_value}
                else:
                    invoice.fields = {}
        else:
            # 複数月モード
            invoice = ocr_service.analyze_invoice(
                content,
                mode="multi",
                start_month=start_month
            )
            # 複数月モードの場合はfieldsに既にデータが入っているはず
        
        return JSONResponse({
            "success": True,
            "filename": file.filename,
            "fields": invoice.fields,
            "raw_text": invoice.raw_text # デバッグ用
        })

    except Exception as e:
        logger.error(f"単一処理エラー: {str(e)}", exc_info=True)
        return JSONResponse({
            "success": False,
            "filename": file.filename,
            "error": str(e)
        }, status_code=500)


@router.post("/generate_excel")
async def generate_excel(request: ExcelGenerationRequest):
    """
    OCR結果のリストを受け取りExcelを生成する
    """
    global _last_excel_path
    
    try:
        cfg = load_app_config()
        excel_service = ExcelService(cfg)
        
        invoices = []
        for data in request.invoices_data:
            # 辞書データからInvoiceオブジェクトを復元（簡易的）
            inv = Invoice(raw_text="")
            inv.fields = data.get("fields", {})
            invoices.append(inv)
            
        excel_path = excel_service.write_invoices(invoices, corp_name=request.corp_name)
        _last_excel_path = excel_path
        
        return JSONResponse({
            "success": True,
            "excel_path": excel_path
        })
        
    except Exception as e:
        logger.error(f"Excel生成エラー: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
